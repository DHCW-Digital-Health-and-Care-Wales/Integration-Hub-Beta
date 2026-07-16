#!/usr/bin/env python3
"""
SQL Stored Procedure Inventory Tool
====================================
Scans a folder of SQL deployment scripts and produces an Excel inventory
of stored procedures with the following columns:

  - Procedure Name      : Full schema.name of the stored procedure
  - Description         : Description extracted from header comments
  - Tables Used         : Comma-separated list of tables referenced
  - Discontinued        : Yes / No
  - Discontinuation Reason : Error/warning message returned, or 'api schema'
  - Replacement Procedure  : Name of the replacement procedure (api schema only)

Usage
-----
  python sql_proc_inventory.py <folder_path> [output.xlsx]

Examples
--------
  python sql_proc_inventory.py "C:\\SQLScripts\\UI_FIMSDeploy"
  python sql_proc_inventory.py "C:\\SQLScripts\\UI_FIMSDeploy" "output\\inventory.xlsx"

Requirements
------------
  pip install openpyxl
"""

import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class StoredProcedure:
    name: str                               # schema.ProcName
    schema: str
    proc_name: str
    description: str
    tables: List[str] = field(default_factory=list)
    is_discontinued: bool = False
    discontinuation_reason: str = ""
    replacement_proc: str = ""
    source_file: str = ""


# ---------------------------------------------------------------------------
# Regex patterns
# ---------------------------------------------------------------------------

# Matches: CREATE [OR ALTER] PROCEDURE [schema].[name] or schema.name
_RE_CREATE_PROC = re.compile(
    r"CREATE\s+(?:OR\s+ALTER\s+)?PROC(?:EDURE)?\s+"
    r"(?:\[?(\w+)\]?\.)?\[?(\w+)\]?",
    re.IGNORECASE,
)

# Table references in DML statements
_RE_TABLE_REF = re.compile(
    r"(?:FROM|JOIN|INTO|UPDATE|MERGE\s+(?:INTO\s+)?|DELETE\s+FROM)\s+"
    r"(?:\[?(\w+)\]?\.)?\[?(\w+)\]?",
    re.IGNORECASE,
)

# RAISERROR / THROW / RETURN with a string message (indicates discontinued body)
_RE_ERROR_MSG = re.compile(
    r"(?:RAISERROR\s*\(\s*['\"](.+?)['\"]|THROW\s+\d+\s*,\s*['\"](.+?)['\"])",
    re.IGNORECASE | re.DOTALL,
)

# Patterns for replacement procedure name in a description
_RE_REPLACEMENT = re.compile(
    r"(?:use|replaced?\s+by|replaced?\s+with|instead\s+use|now\s+use|"
    r"please\s+use|migrate\s+to|see|refer\s+to)\s+"
    r"(?:\[?(\w+)\]?\.\[?(\w+)\]?|\[?(\w+)\]?)",
    re.IGNORECASE,
)

# System / metadata tables to exclude from the tables list
_SYSTEM_TABLES = {
    "sys", "information_schema", "sysobjects", "syscolumns",
    "sysindexes", "sysusers", "master", "tempdb",
}

# SQL keywords that look like table names but aren't
_SQL_KEYWORDS = {
    "select", "where", "set", "values", "default", "null", "not",
    "and", "or", "as", "on", "by", "with", "top", "distinct",
    "inserted", "deleted", "output", "pivot", "unpivot",
}


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def _strip_single_line_comments(sql: str) -> str:
    """Remove -- comments (preserving newlines)."""
    return re.sub(r"--[^\n]*", "", sql)


def _clean_description(text: str) -> str:
    """Remove SQL comment decorators (separator lines of = or - or * characters)."""
    lines = text.splitlines()
    cleaned = [
        ln for ln in lines
        if not re.match(r"^\s*[=\-\*]{4,}\s*$", ln)
    ]
    return "\n".join(cleaned).strip()


def _extract_header_comment(proc_body: str, preceding_text: str = "") -> str:
    """
    Extract description from the header of a stored procedure.

    Looks for comments in this priority order:
      1. Block comment /* ... */ immediately preceding the CREATE PROCEDURE line
         (in the ``preceding_text`` slice of the file)
      2. Consecutive -- comment lines immediately preceding CREATE PROCEDURE
      3. A block comment inside the procedure header (between CREATE and AS/BEGIN)
      4. Consecutive -- lines inside the procedure header

    Returns the cleaned description text.
    """
    # ── 1. Block comment immediately before CREATE PROC ─────────────────────
    # Take up to 3000 chars before the CREATE keyword (covers multi-line headers)
    preceding = preceding_text[-3000:] if len(preceding_text) > 3000 else preceding_text

    # Find the LAST block comment in the preceding section
    block_matches = list(re.finditer(r"/\*(.*?)\*/", preceding, re.DOTALL))
    if block_matches:
        last_block = block_matches[-1]
        # Only use it if there is no significant SQL between it and our CREATE PROC
        gap = preceding[last_block.end():]
        if not re.search(r"\b(?:CREATE|ALTER|DROP|EXEC|EXECUTE)\b", gap, re.IGNORECASE):
            text = last_block.group(1)
            lines = [re.sub(r"^\s*\*?\s?", "", ln) for ln in text.splitlines()]
            return _clean_description("\n".join(ln for ln in lines if ln.strip()).strip())

    # ── 2. Consecutive -- lines immediately before CREATE PROC ──────────────
    # Look for a contiguous block of -- lines ending just before the CREATE
    dash_block = re.findall(r"(?:^|\n)((?:[ \t]*--[^\n]*\n)+)", preceding)
    if dash_block:
        last_dash = dash_block[-1]
        # Check the gap between this comment block and CREATE PROC contains nothing significant
        gap_start = preceding.rfind(last_dash)
        gap = preceding[gap_start + len(last_dash):]
        if not re.search(r"\b(?:CREATE|ALTER|DROP|EXEC|EXECUTE)\b", gap, re.IGNORECASE):
            dash_lines = re.findall(r"--\s?(.*)", last_dash)
            if dash_lines:
                return _clean_description("\n".join(ln.strip() for ln in dash_lines if ln.strip()))

    # ── 3 & 4. Fallback: look inside the proc header (between CREATE and AS) ─
    as_match = re.search(r"\bAS\b|\bBEGIN\b", proc_body, re.IGNORECASE)
    header = proc_body[: as_match.start()] if as_match else proc_body[:500]

    block = re.search(r"/\*(.*?)\*/", header, re.DOTALL)
    if block:
        text = block.group(1)
        lines = [re.sub(r"^\s*\*?\s?", "", ln) for ln in text.splitlines()]
        return _clean_description("\n".join(ln for ln in lines if ln.strip()).strip())

    dash_lines = re.findall(r"--\s?(.*)", header)
    if dash_lines:
        return _clean_description("\n".join(ln.strip() for ln in dash_lines if ln.strip()))

    return ""


def _extract_tables(sql_body: str) -> List[str]:
    """
    Extract table names referenced in DML statements within the procedure body.
    Excludes system tables and SQL keywords.
    """
    tables: Dict[str, str] = {}
    for match in _RE_TABLE_REF.finditer(sql_body):
        schema_part = (match.group(1) or "").lower()
        table_part = (match.group(2) or "").lower()

        if schema_part in _SYSTEM_TABLES or table_part in _SYSTEM_TABLES:
            continue
        if table_part in _SQL_KEYWORDS:
            continue
        if table_part.startswith("#"):  # temp tables
            continue

        full = f"{schema_part}.{table_part}" if schema_part else table_part
        tables[full] = full

    return sorted(tables.values())


def _is_body_commented_out(proc_body: str) -> Tuple[bool, str]:
    """
    Detect whether the procedure body is mostly commented out and returns
    a RAISERROR/THROW error or warning message.

    Returns (is_discontinued, message).
    """
    # Find the AS / BEGIN boundary
    as_match = re.search(r"\bAS\b|\bBEGIN\b", proc_body, re.IGNORECASE)
    if not as_match:
        return False, ""

    body = proc_body[as_match.end():]

    # Check for RAISERROR/THROW (uncommented — active error return)
    err_match = _RE_ERROR_MSG.search(body)
    if not err_match:
        return False, ""

    message = (err_match.group(1) or err_match.group(2) or "").strip()

    # Now determine if the *original* logic is commented out:
    # Strip the RAISERROR/THROW and any surrounding RETURN/SET statements,
    # then check if the majority of the remaining non-trivial content is
    # inside block comments.
    body_no_err = _RE_ERROR_MSG.sub("", body)
    body_no_err = re.sub(r"\bRETURN\b", "", body_no_err, flags=re.IGNORECASE)
    body_no_err = re.sub(r"\bSET\s+\w+\s*=\s*\w+", "", body_no_err, flags=re.IGNORECASE)

    # Extract block comment content length vs total non-whitespace
    block_comment_chars = sum(
        len(m.group(0)) for m in re.finditer(r"/\*.*?\*/", body_no_err, re.DOTALL)
    )
    total_chars = len(re.sub(r"\s+", "", body_no_err))

    if total_chars > 0 and block_comment_chars / total_chars > 0.4:
        return True, message

    # Also treat as discontinued if there's almost no uncommented code
    uncommented = re.sub(r"/\*.*?\*/", "", body_no_err, flags=re.DOTALL)
    uncommented = re.sub(r"--[^\n]*", "", uncommented)
    uncommented_code = re.sub(r"\s+", "", uncommented)

    # If less than 50 chars of actual code remain after removing comments → discontinued
    if len(uncommented_code) < 50 and message:
        return True, message

    return False, ""


def _extract_replacement(description: str, proc_body: str, schema: str) -> str:
    """
    Extract the replacement procedure name from:
    1. The description/header comment
    2. The procedure body (RAISERROR/THROW message or comments)

    For ``api`` schema procedures the replacement is always in a different schema.
    """
    sources = [description, proc_body]

    for source in sources:
        m = _RE_REPLACEMENT.search(source)
        if m:
            s = m.group(1) or ""
            p = m.group(2) or m.group(3) or ""
            if p:
                full = f"{s}.{p}" if s else p
                # Skip if it refers to the same schema (likely self-reference)
                if s.lower() == schema.lower():
                    continue
                return full

    # Fallback: look for any [other_schema].[ProcName] bracket reference in description
    bracket_refs = re.findall(
        r"\[(\w+)\]\.\[(\w+)\]", description + " " + proc_body
    )
    for s, p in bracket_refs:
        if s.lower() != schema.lower():  # different schema = likely a replacement
            return f"[{s}].[{p}]"

    return ""


# ---------------------------------------------------------------------------
# Main SQL file parser
# ---------------------------------------------------------------------------

def parse_sql_file(file_path: Path) -> List[StoredProcedure]:
    """Parse a single SQL file and return all stored procedures found."""
    try:
        content = file_path.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        print(f"  WARNING: Could not read {file_path}: {exc}")
        return []

    procs: List[StoredProcedure] = []

    # Split on CREATE [OR ALTER] PROCEDURE to find each procedure definition
    splits = list(_RE_CREATE_PROC.finditer(content))
    if not splits:
        return []

    for idx, match in enumerate(splits):
        schema = (match.group(1) or "dbo").strip("[]")
        proc_name_raw = (match.group(2) or "").strip("[]")

        # Body is from this match to the next CREATE PROC (or end of file)
        start = match.start()
        end = splits[idx + 1].start() if idx + 1 < len(splits) else len(content)
        proc_body = content[start:end]

        # Text before this procedure (for finding preceding header comments)
        preceding_text = content[:start]

        # Description from header (look before CREATE PROC first, then inside header)
        description = _extract_header_comment(proc_body, preceding_text)

        # Tables used (strip comments first so we don't match inside them)
        body_for_tables = re.sub(r"/\*.*?\*/", " ", proc_body, flags=re.DOTALL)
        body_for_tables = re.sub(r"--[^\n]*", " ", body_for_tables)
        tables = _extract_tables(body_for_tables)

        # Discontinued detection
        is_api_schema = schema.lower() == "api"
        body_commented, err_message = _is_body_commented_out(proc_body)
        is_discontinued = is_api_schema or body_commented

        discontinuation_reason = ""
        if is_api_schema:
            discontinuation_reason = "api schema (deprecated)"
        elif body_commented and err_message:
            discontinuation_reason = err_message

        # Replacement procedure
        replacement = ""
        if is_discontinued:
            replacement = _extract_replacement(description, proc_body, schema)

        procs.append(StoredProcedure(
            name=f"{schema}.{proc_name_raw}",
            schema=schema,
            proc_name=proc_name_raw,
            description=description,
            tables=tables,
            is_discontinued=is_discontinued,
            discontinuation_reason=discontinuation_reason,
            replacement_proc=replacement,
            source_file=file_path.name,
        ))

    return procs


def scan_folder(folder: Path) -> List[StoredProcedure]:
    """Recursively scan a folder for .sql files and extract stored procedures."""
    all_procs: List[StoredProcedure] = []
    sql_files = sorted(folder.rglob("*.sql"))

    if not sql_files:
        print(f"No .sql files found in: {folder}")
        return []

    print(f"Found {len(sql_files)} SQL file(s). Parsing...")
    for sql_file in sql_files:
        print(f"  Processing: {sql_file.name}")
        procs = parse_sql_file(sql_file)
        if procs:
            print(f"    Found {len(procs)} procedure(s)")
        all_procs.extend(procs)

    return all_procs


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

# Colour scheme: DHCW NHS Wales branding
_COLOUR_HEADER_BG = "325083"   # NHS Wales Blue
_COLOUR_HEADER_FG = "FFFFFF"
_COLOUR_DISCONTINUED_BG = "FFDDE1"   # light red
_COLOUR_ACTIVE_BG = "DFF0D8"         # light green
_COLOUR_ALT_ROW_BG = "F5F5F5"


def write_excel(procs: List[StoredProcedure], output_path: Path) -> None:
    """Write the list of stored procedures to an Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stored Procedure Inventory"

    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    headers = [
        "Procedure Name",
        "Description",
        "Tables Used",
        "Discontinued",
        "Discontinuation Reason",
        "Replacement Procedure",
        "Source File",
    ]

    header_fill = PatternFill("solid", fgColor=_COLOUR_HEADER_BG)
    header_font = Font(bold=True, color=_COLOUR_HEADER_FG, name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    discontinued_fill = PatternFill("solid", fgColor=_COLOUR_DISCONTINUED_BG)
    active_fill = PatternFill("solid", fgColor=_COLOUR_ACTIVE_BG)
    alt_fill = PatternFill("solid", fgColor=_COLOUR_ALT_ROW_BG)
    no_fill = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, proc in enumerate(procs, start=2):
        values = [
            proc.name,
            proc.description,
            ", ".join(proc.tables) if proc.tables else "",
            "Yes" if proc.is_discontinued else "No",
            proc.discontinuation_reason,
            proc.replacement_proc,
            proc.source_file,
        ]

        row_fill = discontinued_fill if proc.is_discontinued else (
            alt_fill if row_idx % 2 == 0 else no_fill
        )

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = left_align
            cell.border = thin_border

            # Colour the "Discontinued" column distinctly
            if col_idx == 4:
                cell.fill = discontinued_fill if proc.is_discontinued else active_fill
                cell.font = Font(
                    bold=True,
                    name="Calibri",
                    size=10,
                    color="CC0000" if proc.is_discontinued else "006600",
                )
                cell.alignment = center_align
            else:
                cell.fill = row_fill

    # ------------------------------------------------------------------
    # Column widths
    # ------------------------------------------------------------------
    col_widths = [40, 60, 50, 15, 60, 40, 30]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Auto-filter on headers
    ws.auto_filter.ref = ws.dimensions

    # ------------------------------------------------------------------
    # Summary sheet
    # ------------------------------------------------------------------
    ws_summary = wb.create_sheet(title="Summary")
    total = len(procs)
    discontinued_count = sum(1 for p in procs if p.is_discontinued)
    active_count = total - discontinued_count

    summary_data = [
        ("Total Procedures", total),
        ("Active (not discontinued)", active_count),
        ("Discontinued", discontinued_count),
        ("api schema (deprecated)", sum(1 for p in procs if p.schema.lower() == "api")),
        ("Body commented out", sum(1 for p in procs if p.is_discontinued and p.schema.lower() != "api")),
        ("With replacement specified", sum(1 for p in procs if p.replacement_proc)),
    ]

    ws_summary.cell(row=1, column=1, value="Metric").fill = header_fill
    ws_summary.cell(row=1, column=1).font = header_font
    ws_summary.cell(row=1, column=2, value="Count").fill = header_fill
    ws_summary.cell(row=1, column=2).font = header_font

    for row_idx, (label, value) in enumerate(summary_data, start=2):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 15

    wb.save(output_path)
    print(f"\nExcel inventory saved to: {output_path}")
    print(f"  Total procedures : {total}")
    print(f"  Active           : {active_count}")
    print(f"  Discontinued     : {discontinued_count}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    folder = Path(sys.argv[1])
    if not folder.is_dir():
        print(f"ERROR: Folder not found: {folder}")
        sys.exit(1)

    output_path = Path(sys.argv[2]) if len(sys.argv) >= 3 else folder / "stored_proc_inventory.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    procs = scan_folder(folder)

    if not procs:
        print("No stored procedures found.")
        sys.exit(0)

    write_excel(procs, output_path)


if __name__ == "__main__":
    main()
