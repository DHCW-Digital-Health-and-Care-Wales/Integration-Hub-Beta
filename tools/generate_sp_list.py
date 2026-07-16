#!/usr/bin/env python3
"""
Generate an Excel inventory of SQL stored procedures from deployment script files.

Scans every .sql file in a given folder (recursively), extracts each stored procedure,
and writes a formatted Excel spreadsheet with the following columns:

    Stored Procedure Name  – fully qualified schema.name
    Schema                 – the schema only
    Description            – extracted from leading SQL comments
    Tables Used            – comma-separated list of tables referenced in the body
    Is Discontinued        – Yes / No
                             Discontinued when:
                               * the schema is 'api' (legacy API layer being retired), OR
                               * the active (non-commented) body contains RAISERROR / THROW
                                 (code has been commented out and an error/warning returned)
    Replacement SP         – for discontinued procedures, the replacement name parsed
                             from the description comments

Usage (requires openpyxl):
    uv run --with openpyxl python tools/generate_sp_list.py <sql_folder> [output.xlsx]
"""

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "openpyxl is required.  Install with:  uv run --with openpyxl python tools/generate_sp_list.py",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------


@dataclass
class StoredProcedure:
    name: str
    schema: str
    full_name: str
    description: str
    tables_used: List[str] = field(default_factory=list)
    is_discontinued: bool = False
    replacement_sp: Optional[str] = None
    source_file: str = ""


# ---------------------------------------------------------------------------
# Compiled regex patterns
# ---------------------------------------------------------------------------

# CREATE [OR ALTER] PROCEDURE  [schema].[name]  (with optional brackets)
_RE_SP_HEADER = re.compile(
    r"CREATE\s+(?:OR\s+ALTER\s+)?PROC(?:EDURE)?\s+"
    r"(?:\[?(\w+)\]?\s*\.\s*)?\[?(\w+)\]?",
    re.IGNORECASE,
)

# Tables in FROM / JOIN / INTO / UPDATE / MERGE / DELETE FROM clauses.
# Group 1+2: schema.table   |   Group 3: bare table name.
# NOTE: function-call filtering (TVFs) is handled in _extract_tables by checking
# whether the character immediately after the match is '(' (no space = function call).
_RE_TABLE_REF = re.compile(
    r"(?:FROM|JOIN|INTO|UPDATE|MERGE(?:\s+INTO)?|DELETE\s+FROM)\s+"
    r"(?:\[?(\w+)\]?\s*\.\s*\[?(\w+)\]?"    # schema.table
    r"|\[?(\w+)\]?)",                         # bare table
    re.IGNORECASE,
)

# RAISERROR or THROW in the *active* (non-commented) body
_RE_RAISE_THROW = re.compile(r"\b(?:RAISERROR|THROW)\b", re.IGNORECASE)

# GO batch separator on its own line
_RE_GO = re.compile(r"^\s*GO\s*$", re.MULTILINE | re.IGNORECASE)

# Patterns that suggest a replacement SP name in free text
_RE_REPLACEMENT = [
    re.compile(
        r"(?:use|replaced?\s+by|replacement[:\s]+|see|migrated?\s+to|please\s+use|instead\s+use)\s+"
        r"[\[`\[]?(\w+(?:\.\w+)+)[\]`\]]?",
        re.IGNORECASE,
    ),
    re.compile(
        r"Replacement\s*[:\-]\s*[\[`\[]?(\w+(?:\.\w+)*)[\]`\]]?",
        re.IGNORECASE,
    ),
    re.compile(
        r"New\s+(?:SP|procedure|proc)\s*[:\-\s]+[\[`\[]?(\w+(?:\.\w+)*)[\]`\]]?",
        re.IGNORECASE,
    ),
]

# SQL reserved words, SQL Server pseudo-tables (inserted/deleted — only available inside
# triggers, never valid as base-table names), and built-in function names to exclude from
# the table list.  These are separated into logical groups for clarity.
_SQL_KEYWORDS = frozenset(
    {
        # DML / DDL keywords
        "select", "from", "where", "set", "exec", "execute", "go", "begin", "end",
        "declare", "print", "if", "else", "while", "return", "raiserror", "throw",
        "nolock", "with", "on", "inner", "outer", "left", "right", "cross", "full",
        "join", "and", "or", "not", "null", "is", "as", "top", "distinct", "into",
        "delete", "update", "insert", "merge", "values", "output", "using", "when",
        "then", "matched", "by",
        # Trigger pseudo-tables (cannot be user-defined base tables in SQL Server)
        "inserted", "deleted",
        # MERGE clause pseudo-aliases
        "target", "source",
        # Common built-in functions that appear after FROM/JOIN in TVF-like usage
        "sysdatetimeoffset", "getdate", "getutcdate", "newid",
    }
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _strip_sql_comments(sql: str) -> str:
    """Remove /* … */ block comments and -- line comments, returning active SQL."""
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)
    sql = re.sub(r"--[^\n]*", " ", sql)
    return sql


def _extract_description(sp_text: str) -> str:
    """
    Pull the human-readable description from comments that appear between
    the CREATE PROCEDURE header and the AS keyword.
    """
    # Take everything before the AS keyword that starts the procedure body
    before_as_match = re.search(r"\bAS\b", sp_text, re.IGNORECASE)
    header_section = sp_text[: before_as_match.start()] if before_as_match else sp_text

    lines: List[str] = []

    # Block comments  /* … */
    for block in re.findall(r"/\*(.*?)\*/", header_section, re.DOTALL):
        for raw_line in block.splitlines():
            text = raw_line.strip().lstrip("*").strip()
            if text and not re.fullmatch(r"[-=*]+", text):
                lines.append(text)

    # Line comments  -- …
    for raw_line in header_section.splitlines():
        stripped = raw_line.strip()
        if stripped.startswith("--"):
            text = stripped.lstrip("-").strip()
            if text and not re.fullmatch(r"[-=*]+", text):
                lines.append(text)

    return " ".join(lines).strip()


def _extract_tables(sp_text: str) -> List[str]:
    """
    Return a sorted, deduplicated list of table names referenced in the SP body.
    Excludes temp tables (#…), table variables (@…), SQL keywords, and TVF calls
    (identifiers immediately followed by '(' with no intervening space).
    """
    tables: set[str] = set()

    for m in _RE_TABLE_REF.finditer(sp_text):
        # Exclude table-valued function calls: '(' immediately after the identifier (no space)
        if sp_text[m.end() : m.end() + 1] == "(":
            continue

        if m.group(1) and m.group(2):
            # schema.table — keep as-is (normalise brackets away)
            table = f"{m.group(1)}.{m.group(2)}"
        else:
            name = (m.group(3) or "").strip()
            if not name or name.startswith(("#", "@")):
                continue
            if name.lower() in _SQL_KEYWORDS:
                continue
            table = name

        # Strip any remaining brackets
        table = table.replace("[", "").replace("]", "")

        # Skip if any component is a keyword
        parts = [p.lower() for p in table.split(".")]
        if any(p in _SQL_KEYWORDS for p in parts):
            continue

        tables.add(table)

    return sorted(tables)


def _is_discontinued(sp_text: str, schema: str) -> bool:
    """
    Return True when the procedure is being discontinued:
      - Schema is 'api' (the entire api schema is being retired), OR
      - The non-commented body contains RAISERROR / THROW, meaning the
        actual logic has been commented out and an error/warning is returned.
    """
    if schema.lower() == "api":
        return True
    active_code = _strip_sql_comments(sp_text)
    return bool(_RE_RAISE_THROW.search(active_code))


def _extract_replacement(description: str, sp_text: str) -> Optional[str]:
    """
    Look for a replacement stored procedure name in the description or
    any comments inside the procedure body.
    """
    sources = [description]

    # Collect body comments (block + line)
    sources += re.findall(r"/\*(.*?)\*/", sp_text, re.DOTALL)
    sources += re.findall(r"--([^\n]*)", sp_text)

    combined = " ".join(sources)
    for pattern in _RE_REPLACEMENT:
        m = pattern.search(combined)
        if m:
            candidate = m.group(1).strip().strip("[]`")
            return candidate
    return None


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


def parse_stored_procedures(sql_content: str, source_file: str) -> List[StoredProcedure]:
    """
    Parse every CREATE [OR ALTER] PROCEDURE statement found in *sql_content*
    and return a list of :class:`StoredProcedure` objects.

    Description comments are extracted from BOTH:
    - The text preceding the CREATE PROCEDURE header in the same GO batch
      (the most common placement for descriptive comment blocks), AND
    - Any comments between the header and the AS keyword.
    """
    results: List[StoredProcedure] = []

    # Find every SP header position
    sp_matches = list(_RE_SP_HEADER.finditer(sql_content))

    for idx, start_m in enumerate(sp_matches):
        # Slice text for this SP: up to the next SP header (or EOF)
        end_pos = sp_matches[idx + 1].start() if idx + 1 < len(sp_matches) else len(sql_content)
        sp_text = sql_content[start_m.start() : end_pos]

        # Trim at the GO batch separator
        go_m = _RE_GO.search(sp_text)
        if go_m:
            sp_text = sp_text[: go_m.start()]

        # Also capture the preceding batch (comments before CREATE PROCEDURE)
        preceding_text = sql_content[: start_m.start()]
        # Find the last GO before this SP (or start of file)
        go_positions = [m.end() for m in _RE_GO.finditer(preceding_text)]
        batch_start = go_positions[-1] if go_positions else 0
        preceding_batch = preceding_text[batch_start:]

        schema = (start_m.group(1) or "dbo").strip().strip("[]")
        name = start_m.group(2).strip().strip("[]")
        full_name = f"{schema}.{name}"

        # Extract description from both preceding batch comments and the SP header section
        description = _extract_description(preceding_batch + sp_text)
        tables = _extract_tables(sp_text)
        discontinued = _is_discontinued(sp_text, schema)
        replacement = _extract_replacement(description, sp_text) if discontinued else None

        results.append(
            StoredProcedure(
                name=name,
                schema=schema,
                full_name=full_name,
                description=description,
                tables_used=tables,
                is_discontinued=discontinued,
                replacement_sp=replacement,
                source_file=source_file,
            )
        )

    return results


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

# DHCW brand colours
_COLOUR_HEADER_BG = "325083"   # NHS Wales Blue
_COLOUR_HEADER_FG = "FFFFFF"
_COLOUR_DISC_BG = "FFD7D7"    # Soft red for discontinued rows
_COLOUR_ACTIVE_BG = "FFFFFF"


def generate_excel(procedures: List[StoredProcedure], output_path: str) -> None:
    """Write *procedures* to a formatted Excel workbook at *output_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stored Procedures"

    # Styles
    header_fill = PatternFill(start_color=_COLOUR_HEADER_BG, end_color=_COLOUR_HEADER_BG, fill_type="solid")
    header_font = Font(color=_COLOUR_HEADER_FG, bold=True, name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    disc_fill = PatternFill(start_color=_COLOUR_DISC_BG, end_color=_COLOUR_DISC_BG, fill_type="solid")
    active_fill = PatternFill(start_color=_COLOUR_ACTIVE_BG, end_color=_COLOUR_ACTIVE_BG, fill_type="solid")
    cell_align = Alignment(vertical="top", wrap_text=True)

    headers = [
        "Stored Procedure Name",
        "Schema",
        "Description",
        "Tables Used",
        "Is Discontinued",
        "Replacement Stored Procedure",
        "Source File",
    ]
    col_widths = [42, 14, 65, 55, 16, 42, 32]

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, sp in enumerate(procedures, start=2):
        row_fill = disc_fill if sp.is_discontinued else active_fill
        values = [
            sp.full_name,
            sp.schema,
            sp.description,
            ", ".join(sp.tables_used),
            "Yes" if sp.is_discontinued else "No",
            sp.replacement_sp or "",
            sp.source_file,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = cell_align

    # Column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row and add auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Saved: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate an Excel inventory of stored procedures from SQL deployment scripts.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Example:\n"
            "  uv run --with openpyxl python tools/generate_sp_list.py ./sql_scripts output.xlsx"
        ),
    )
    parser.add_argument("sql_folder", help="Folder containing .sql deployment script files (searched recursively)")
    parser.add_argument(
        "output",
        nargs="?",
        default="stored_procedures.xlsx",
        help="Path for the output Excel file (default: stored_procedures.xlsx)",
    )
    args = parser.parse_args()

    sql_folder = Path(args.sql_folder)
    if not sql_folder.is_dir():
        print(f"Error: '{sql_folder}' is not a directory.", file=sys.stderr)
        return 1

    sql_files = sorted(sql_folder.rglob("*.sql"))
    if not sql_files:
        print(f"No .sql files found in '{sql_folder}'.", file=sys.stderr)
        return 1

    all_procedures: List[StoredProcedure] = []
    for sql_file in sql_files:
        try:
            content = sql_file.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            print(f"  Warning: could not read {sql_file.name}: {exc}", file=sys.stderr)
            continue
        try:
            procs = parse_stored_procedures(content, str(sql_file.relative_to(sql_folder)))
            all_procedures.extend(procs)
            print(f"  {sql_file.name}: {len(procs)} stored procedure(s)")
        except (ValueError, re.error) as exc:
            print(f"  Warning: could not parse {sql_file.name}: {exc}", file=sys.stderr)

    if not all_procedures:
        print("No stored procedures were found across the SQL files.")
        return 0

    print(f"\nTotal: {len(all_procedures)} stored procedure(s) found")
    generate_excel(all_procedures, args.output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
